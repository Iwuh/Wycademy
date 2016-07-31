import openpyxl
import collections
import json

wb = openpyxl.load_workbook(r"Monster Info.xlsx")

# There are 71 sheets in the workbook, 1 for each monster
for i in wb.sheetnames:
    # Reset dictionaries for each sheet
    hitzones = collections.OrderedDict()
    staggersever = collections.OrderedDict()
    status = collections.OrderedDict()
    current_sheet = wb.get_sheet_by_name(i)
    # add columns B & C as a key:value pair to a dictionary depending on which category is in A
    for j in range(1, current_sheet.max_row + 1):
        if current_sheet["A{0}".format(j)].value == "Hitzone":
            hitzones[current_sheet["B{0}".format(j)].value] = current_sheet["C{0}".format(j)].value
        elif current_sheet["A{0}".format(j)].value == "Stagger/Sever":
            staggersever[current_sheet["B{0}".format(j)].value] = current_sheet["C{0}".format(j)].value
        elif current_sheet["A{0}".format(j)].value == "Status":
            status[current_sheet["B{0}".format(j)].value] = current_sheet["C{0}".format(j)].value
        elif current_sheet["A{0}".format(j)].value == "Item Effects":
            status[current_sheet["B{0}".format(j)].value] = current_sheet["C{0}".format(j)].value
        else:
            # If none of the previous matched then raise an error because that shouldn't happen and it needs to be fixed
            raise ValueError("Unknown category: " + current_sheet["A{0}".format(j)].value)

    # write the data to a json file, prettyprinted with 4 space indentations
    with open("json\{0}.json".format(i), "w+") as f:
        json.dump({'hitzones': hitzones, 'stagger/sever': staggersever, 'status': status},
                  f, indent=4)
