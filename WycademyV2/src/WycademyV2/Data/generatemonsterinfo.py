import openpyxl
from collections import OrderedDict
from json import dump


wb = openpyxl.load_workbook(r"Monster Info.xlsx")

for i in wb.sheetnames:
    hitzones = OrderedDict()
    staggersever = OrderedDict()
    status = OrderedDict()
    itemeffects = OrderedDict()

    current_sheet = wb.get_sheet_by_name(i)

    for row in range(1, current_sheet.max_row + 1):
        category = current_sheet["A%s" % row].value
        key = current_sheet["B%s" % row].value
        value = current_sheet["C%s" % row].value

        if category == "Hitzone":
            hitzones[key] = value
        elif category == "Stagger/Sever":
            staggersever[key] = value
        elif category == "Status":
            status[key] = value
        elif category == "Item Effects":
            itemeffects[key] = value
        else:
            raise ValueError("Invalid category")

        with open("./monster/%s.json" % i, "w+", encoding="utf-8") as f:
            dump({"Hitzone": hitzones, "Stagger/Sever": staggersever, "Status": status, "Item Effects": itemeffects},
                 f, indent=4)
